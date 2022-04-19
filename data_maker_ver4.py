from ast import arguments
from ctypes import alignment
from operator import mod
import os
import numpy as np
from sqlite3 import Row
from numpy import require
import pandas as pd
import argparse
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import sys
import shutil


TS = "出纳账【总帐-东京星银行】"
R = "出纳账【总帐-乐天银行】"
M = "出纳账【总帐-瑞穗银行】 "
BC = "出纳账【总帐-中国银行】"
Original_data_frame_Column_Names = ["报销单摘要", "支付金额", "附签"]
Original_data_frame = pd.DataFrame(columns=Original_data_frame_Column_Names)


def OutPutmaker(input_data_frame, outputfile_path, tempfile_path, TargetYear, TargetMonth, ExchangeRate):
    # Copy the temp excel book to the output_folder, and rename the book name
    # original_excel = tempfile_path
    target_excel_path = outputfile_path + "/" + \
        TargetYear + "年_" + TargetMonth + "月报销单.xlsx"
    # 生成目标Excel
    shutil.copyfile(tempfile_path, target_excel_path)
    wb_target_excel_wb = openpyxl.load_workbook(target_excel_path)
    sheet_name_in_wb_target_excel = ''
    ws = ''
    print(f'len = {len(input_data_frame)}')
    for i in range(len(input_data_frame)):
        print(f'i = {i}')
        # Define the input
        # Define the input cell location
        报销单摘要 = "A19"
        支付金额 = "D19"
        附签 = "B14"
        汇率 = "G20"
        折合人名币 = "F19"
        copy_flag = "do_not_copy_template_sheet"
        if i % 2 == 0:
            sheet_name_in_wb_target_excel = str(i+1) + "," + str(i+2)
            # wb_target_excel_wb.create_sheet(sheet_name_in_wb_target_excel)
            # newly define the input cell location
            报销单摘要 = "A7"
            支付金额 = "D7"
            附签 = "B2"
            汇率 = "G8"
            折合人名币 = "F7"
            copy_flag = "copy_template_sheet"
            print(f'i = {i}')
            wb_target_excel_wb.active = int((i/2)+1)
            # print(f'sheet_name = {wb_target_excel_wb.active}')
            # ws = wb_target_excel_wb.active
            # in the target_excel, copy the sheet "Template_请不要修改" to another new one, and name the sheet.
        if copy_flag == "copy_template_sheet":
            worksheet = wb_target_excel_wb.copy_worksheet(
                wb_target_excel_wb["Template_请不要修改"])
            worksheet.title = sheet_name_in_wb_target_excel
            worksheet[报销单摘要].value = input_data_frame.iloc[i, 0]
            worksheet[支付金额].value = int(input_data_frame.iloc[i, 1])
            worksheet[附签].value = input_data_frame.iloc[i, 2]
            worksheet[汇率].value = ExchangeRate
            worksheet[折合人名币].value = '\u00A5' + \
                str(int(worksheet[支付金额].value * float(worksheet[汇率].value)))
            worksheet[支付金额] = '\u00A5' + str(input_data_frame.iloc[i, 1])
            wb_target_excel_wb.save(target_excel_path)
            wb_target_excel_wb.close()
        else:
            worksheet[报销单摘要] = input_data_frame.iloc[i, 0]
            worksheet[支付金额] = int(input_data_frame.iloc[i, 1])
            worksheet[附签] = input_data_frame.iloc[i, 2]
            worksheet[汇率] = ExchangeRate
            worksheet[折合人名币] = '\u00A5' + \
                str(int(worksheet[支付金额].value * float(worksheet[汇率].value)))
            worksheet[支付金额] = '\u00A5' + str(input_data_frame.iloc[i, 1])
            wb_target_excel_wb.save(target_excel_path)
            wb_target_excel_wb.close()
        # return target_excel_path


def Data_Processor(data_frame, TargetYear, TargetMonth):
    # drop all rows with NaN in data_frame
    temp_data_frame = data_frame.dropna()
    mod_data_frame = pd.DataFrame()
    for i in range(len(temp_data_frame)):
        if TargetYear in temp_data_frame.iloc[i, 2]:
            # check if TargetMonth exist
            if TargetMonth == temp_data_frame.iloc[i, 2].split('-')[1]:
                mod_data_frame = mod_data_frame.append(temp_data_frame.iloc[i])
        else:
            continue
    return mod_data_frame


def Data_Frame_Generator(Inputfile_Path, Sheet_Name):
    Original_data_frame = pd.read_excel(
        Inputfile_Path, Sheet_Name, usecols="G,I,K")
    return Original_data_frame


def UserInputParser():
    parser = argparse.ArgumentParser('Parsering the I/O')
    parser.add_argument('--inputfile', '-i',
                        dest="inputfile", default='c:\\Users\\rakou\\Tools\\Data_maker\\input\\星辰账本2021.xlsx', type=str, required=True)
    parser.add_argument('--outputfolder', '-o',
                        dest="outputfolder", default='c:\\Users\\rakou\\Tools\\Data_maker\\output', type=str, required=True)
    parser.add_argument('--tempfile', '-t', dest="tempfile", default='c:\\Users\\rakou\\Tools\\Data_maker\\temp\\out_put_temp.xlsx',
                        type=str, required=True)
    parser.add_argument('--ExchangeRate', '-e',
                        dest="ExchangeRate", type=float, required=True)
    parser.add_argument(
        '--Year', '-y', dest="Year", type=str, required=True)
    parser.add_argument('--Month', '-m', dest="Month", type=str, required=True)
    parser.add_argument('--SheetName', '-s',
                        dest="Input_SheetName", type=str, required=True)
    arguments = parser.parse_args()
    Inputfile_path = arguments.inputfile
    Outputfile_path = arguments.outputfolder
    Tempfile_path = arguments.tempfile
    Inputfile_path = os.path.abspath(Inputfile_path)
    Outputfile_path = os.path.abspath(Outputfile_path)
    Tempfile_path = os.path.abspath(Tempfile_path)
    # ExchangeRate = arguments.ExchangeRate
    # Year = arguments.Year
    # Month = arguments.Month
    # Input_SheetName = arguments.SheetName
    return arguments


def main(argv):
    UserInputParser()
    Inputfile_path = UserInputParser().inputfile
    # print(f"Inputfile_path = ")
    # Inputfile_path = 'c:\\Users\\rakou\\Tools\\Data_maker\\input\\星辰账本2021.xlsx'
    Outputfile_path = UserInputParser().outputfolder
    # Outputfile_path = 'c:\\Users\\rakou\\Tools\\Data_maker\\output'
    Tempfile_path = UserInputParser().tempfile
    # Tempfile_path = 'c:\\Users\\rakou\\Tools\\Data_maker\\temp\\out_put_temp.xlsx'
    ExchangeRate = UserInputParser().ExchangeRate
    Year = UserInputParser().Year
    Month = UserInputParser().Month
    Input_Sheet_Name = UserInputParser().Input_SheetName
    # print(f"InputFile_path_in main = {Inputfile_path}")
    Original_data_frame = Data_Frame_Generator(
        Inputfile_path, Input_Sheet_Name)
    Modified_data_frame = Data_Processor(Original_data_frame, Year, Month)
    # print(f'Modified data frame = {Modified_data_frame}')
    OutPutmaker(Modified_data_frame, Outputfile_path,
                Tempfile_path, Year, Month, ExchangeRate)


if __name__ == "__main__":
    main(sys.argv)
