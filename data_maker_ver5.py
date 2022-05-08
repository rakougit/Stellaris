from ast import arguments
from ctypes import alignment
from operator import mod
import os
import numpy as np
from sqlite3 import Row
from numpy import require
import pandas as pd
import argparse
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import sys
import shutil
import num_to_hanzi_copied
import warnings
from decimal import Decimal


TS = "出纳账【总帐-东京星银行】"
R = "出纳账【总帐-乐天银行】"
M = "出纳账【总帐-瑞穗银行】 "
BC = "出纳账【总帐-中国银行】"
Original_data_frame_Column_Names = ["报销单摘要", "支付金额", "附签"]
Original_data_frame = pd.DataFrame(columns=Original_data_frame_Column_Names)

def cncurrency(value, capital=True, prefix=False, classical=None):
    '''
    参数:
    capital:    True   大写汉字金额
                False  一般汉字金额
    classical:  True   元
                False  圆
    prefix:     True   以'人民币'开头
                False, 无开头
    '''
    if not isinstance(value, (Decimal, str, int)):
        msg = '''
        由于浮点数精度问题，请考虑使用字符串，或者 decimal.Decimal 类。
        因使用浮点数造成误差而带来的可能风险和损失作者概不负责。
        '''
        warnings.warn(msg, UserWarning)
    # 默认大写金额用圆，一般汉字金额用元
    if classical is None:
        classical = True if capital else False

    # 汉字金额前缀
    if prefix is True:
        prefix = '人民币'
    else:
        prefix = ''

    # 汉字金额字符定义
    dunit = ('角', '分')
    if capital:
        num = ('零', '壹', '贰', '叁', '肆', '伍', '陆', '柒', '捌', '玖')
        iunit = [None, '拾', '佰', '仟', '万', '拾', '佰', '仟','亿', '拾', '佰', '仟', '万', '拾', '佰', '仟']
    else:
        num = ('〇', '一', '二', '三', '四', '五', '六', '七', '八', '九')
        iunit = [None, '十', '百', '千', '万', '十', '百', '千','亿', '十', '百', '千', '万', '十', '百', '千']
    if classical:
        iunit[0] = '元' if classical else '圆'
    # 转换为Decimal，并截断多余小数

    if not isinstance(value, Decimal):
        value = Decimal(value).quantize(Decimal('0.01'))

    # 处理负数
    if value < 0:
        prefix += '负'          # 输出前缀，加负
        value = - value         # 取正数部分，无须过多考虑正负数舍入
                                # assert - value + value == 0
    # 转化为字符串
    s = str(value)
    if len(s) > 19:
        raise ValueError('金额太大了，不知道该怎么表达。')
    istr, dstr = s.split('.')           # 小数部分和整数部分分别处理
    istr = istr[::-1]                   # 翻转整数部分字符串
    so = []     # 用于记录转换结果

    # 零
    if value == 0:
        return prefix + num[0] + iunit[0]
    haszero = False     # 用于标记零的使用
    if dstr == '00':
        haszero = True  # 如果无小数部分，则标记加过零，避免出现“圆零整”

    # 处理小数部分
    # 分
    if dstr[1] != '0':
        so.append(dunit[1])
        so.append(num[int(dstr[1])])
    else:
        so.append('整')         # 无分，则加“整”
    # 角
    if dstr[0] != '0':
        so.append(dunit[0])
        so.append(num[int(dstr[0])])
    elif dstr[1] != '0':
        so.append(num[0])       # 无角有分，添加“零”
        haszero = True          # 标记加过零了

    # 无整数部分
    if istr == '0':
        if haszero:             # 既然无整数部分，那么去掉角位置上的零
            so.pop()
        so.append(prefix)       # 加前缀
        so.reverse()            # 翻转
        return ''.join(so)

    # 处理整数部分
    for i, n in enumerate(istr):
        n = int(n)
        if i % 4 == 0:          # 在圆、万、亿等位上，即使是零，也必须有单位
            if i == 8 and so[-1] == iunit[4]:   # 亿和万之间全部为零的情况
                so.pop()                        # 去掉万
            so.append(iunit[i])
            if n == 0:                          # 处理这些位上为零的情况
                if not haszero:                 # 如果以前没有加过零
                    so.insert(-1, num[0])       # 则在单位后面加零
                    haszero = True              # 标记加过零了
            else:                               # 处理不为零的情况
                so.append(num[n])
                haszero = False                 # 重新开始标记加零的情况
        else:                                   # 在其他位置上
            if n != 0:                          # 不为零的情况
                so.append(iunit[i])
                so.append(num[n])
                haszero = False                 # 重新开始标记加零的情况
            else:                               # 处理为零的情况
                if not haszero:                 # 如果以前没有加过零
                    so.append(num[0])
                    haszero = True

    # 最终结果
    so.append(prefix)
    so.reverse()
    return ''.join(so)



def OutPutmaker(input_data_frame, outputfile_path, tempfile_path, TargetYear, TargetMonth, ExchangeRate):
    # Copy the temp excel book to the output_folder, and rename the book name
    # original_excel = tempfile_path
    target_excel_path = outputfile_path + "/" + \
        TargetYear + "年_" + TargetMonth + "月报销单.xlsx"
    # 生成目标Excel
    shutil.copyfile(tempfile_path, target_excel_path)
    wb_target_excel_wb = openpyxl.load_workbook(target_excel_path)
    sheet_name_in_wb_target_excel = ''
    ws = ''
    # print(f'len = {len(input_data_frame)}')
    for i in range(len(input_data_frame)):
        # print(f'i = {i}')
        # Define the input
        # Define the input cell location
        报销单摘要 = "A19"
        支付金额 = "D19"
        附签 = "B14"
        汇率 = "G20"
        折合人名币 = "F19"
        数字转汉字 = "C21"
        copy_flag = "do_not_copy_template_sheet"
        if i % 2 == 0:
            sheet_name_in_wb_target_excel = str(i+1) + "," + str(i+2)
            # wb_target_excel_wb.create_sheet(sheet_name_in_wb_target_excel)
            # newly define the input cell location
            报销单摘要 = "A7"
            支付金额 = "D7"
            附签 = "B2"
            汇率 = "G8"
            折合人名币 = "F7"
            数字转汉字 = "C9"
            copy_flag = "copy_template_sheet"
            print(f'i = {i}')
            wb_target_excel_wb.active = int((i/2)+1)
            # print(f'sheet_name = {wb_target_excel_wb.active}')
            # ws = wb_target_excel_wb.active
            # in the target_excel, copy the sheet "Template_请不要修改" to another new one, and name the sheet.
        if copy_flag == "copy_template_sheet":
            worksheet = wb_target_excel_wb.copy_worksheet(
                wb_target_excel_wb["Template_请不要修改"])
            worksheet.title = sheet_name_in_wb_target_excel
            worksheet[报销单摘要].value = input_data_frame.iloc[i, 0]
            worksheet[支付金额].value = int(input_data_frame.iloc[i, 1])
            worksheet[附签].value = input_data_frame.iloc[i, 2]
            worksheet[汇率].value = ExchangeRate
            worksheet[折合人名币].value = '\u00A5' + \
                str(int(worksheet[支付金额].value * float(worksheet[汇率].value)))
            worksheet[支付金额] = '\u00A5' + str(input_data_frame.iloc[i, 1])
            worksheet[数字转汉字] = cncurrency(worksheet[支付金额].value, capital=True, prefix=False, classical=None)
            wb_target_excel_wb.save(target_excel_path)
            wb_target_excel_wb.close()
        else:
            worksheet[报销单摘要] = input_data_frame.iloc[i, 0]
            worksheet[支付金额] = int(input_data_frame.iloc[i, 1])
            worksheet[附签] = input_data_frame.iloc[i, 2]
            worksheet[汇率] = ExchangeRate
            worksheet[折合人名币] = '\u00A5' + \
                str(int(worksheet[支付金额].value * float(worksheet[汇率].value)))
            worksheet[支付金额] = '\u00A5' + str(input_data_frame.iloc[i, 1])
            worksheet[数字转汉字] = cncurrency(worksheet[支付金额].value, capital=True, prefix=False, classical=None)
            wb_target_excel_wb.save(target_excel_path)
            wb_target_excel_wb.close()
        # return target_excel_path


def Data_Processor(data_frame, TargetYear, TargetMonth):
    # drop all rows with NaN in data_frame
    temp_data_frame = data_frame.dropna()
    mod_data_frame = pd.DataFrame()
    for i in range(len(temp_data_frame)):
        if TargetYear in temp_data_frame.iloc[i, 2]:
            # check if TargetMonth exist
            if TargetMonth == temp_data_frame.iloc[i, 2].split('-')[1]:
                mod_data_frame = mod_data_frame.append(temp_data_frame.iloc[i])
        else:
            continue
    return mod_data_frame


def Data_Frame_Generator(Inputfile_Path, Sheet_Name):
    Original_data_frame = pd.read_excel(
        Inputfile_Path, Sheet_Name, usecols="J,N,Q")
    return Original_data_frame


def UserInputParser():
    parser = argparse.ArgumentParser('Parsering the I/O')
    parser.add_argument('--inputfile', '-i',
                        dest="inputfile", default='c:\\Users\\rakou\\Tools\\Data_maker\\input\\星辰账本2021.xlsx', type=str, required=True)
    parser.add_argument('--outputfolder', '-o',
                        dest="outputfolder", default='c:\\Users\\rakou\\Tools\\Data_maker\\output', type=str, required=True)
    parser.add_argument('--tempfile', '-t', dest="tempfile", default='c:\\Users\\rakou\\Tools\\Data_maker\\temp\\out_put_temp.xlsx',
                        type=str, required=True)
    parser.add_argument('--ExchangeRate', '-e',
                        dest="ExchangeRate", type=float, required=True)
    parser.add_argument(
        '--Year', '-y', dest="Year", type=str, required=True)
    parser.add_argument('--Month', '-m', dest="Month", type=str, required=True)
    parser.add_argument('--SheetName', '-s',
                        dest="Input_SheetName", type=str, required=True)
    arguments = parser.parse_args()
    Inputfile_path = arguments.inputfile
    Outputfile_path = arguments.outputfolder
    Tempfile_path = arguments.tempfile
    Inputfile_path = os.path.abspath(Inputfile_path)
    Outputfile_path = os.path.abspath(Outputfile_path)
    Tempfile_path = os.path.abspath(Tempfile_path)
    # ExchangeRate = arguments.ExchangeRate
    # Year = arguments.Year
    # Month = arguments.Month
    # Input_SheetName = arguments.SheetName
    return arguments


def main(argv):
    UserInputParser()
    Inputfile_path = UserInputParser().inputfile
    # print(f"Inputfile_path = ")
    # Inputfile_path = 'c:\\Users\\rakou\\Tools\\Data_maker\\input\\星辰账本2021.xlsx'
    Outputfile_path = UserInputParser().outputfolder
    # Outputfile_path = 'c:\\Users\\rakou\\Tools\\Data_maker\\output'
    Tempfile_path = UserInputParser().tempfile
    # Tempfile_path = 'c:\\Users\\rakou\\Tools\\Data_maker\\temp\\out_put_temp.xlsx'
    ExchangeRate = UserInputParser().ExchangeRate
    Year = UserInputParser().Year
    Month = UserInputParser().Month
    Input_Sheet_Name = UserInputParser().Input_SheetName
    # print(f"InputFile_path_in main = {Inputfile_path}")
    Original_data_frame = Data_Frame_Generator(
        Inputfile_path, Input_Sheet_Name)
    Modified_data_frame = Data_Processor(Original_data_frame, Year, Month)
    # print(f'Modified data frame = {Modified_data_frame}')
    OutPutmaker(Modified_data_frame, Outputfile_path,
                Tempfile_path, Year, Month, ExchangeRate)


if __name__ == "__main__":
    main(sys.argv)
